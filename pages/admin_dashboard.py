import streamlit as st
import pandas as pd
import credentials as cr
from openpyxl import load_workbook
import os

st.set_page_config(layout="wide")
hide_sidebar_style = """
           <style>
               [data-testid="stSidebar"] {
                   display: none;
               }
           </style>
       """
st.markdown(hide_sidebar_style, unsafe_allow_html=True)

if  cr.key == 2:
    st.success("Welcome to dashboard")

    def save_data_to_excel(name, choice):
        file_name = './user_data1.xlsx'
        data = pd.DataFrame([[name, choice]], columns=['Name', 'Choice'])

        if not os.path.exists(file_name):
            # If the file doesn't exist, create it
            data.to_excel(file_name, index=False)
        else:
            try:

                existing_data = pd.read_excel(file_name)
                duplicate = (existing_data['Name'] == name) & (existing_data['Choice'] == choice)

                if duplicate.any():
                    st.warning(f"Duplicate entry found: {name}, {choice}")
                else:
                    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # Getting the last row in the existing sheet
                        book = load_workbook(file_name)
                        sheet = book.active
                        startrow = sheet.max_row

                        data.to_excel(writer, index=False, header=False, startrow=startrow)
                        st.success(f"Data saved: {name}, {choice}")
            except Exception as e:
                st.error(f"Error: {e}")

    def display(filename='user_data1.xlsx'):
        if os.path.exists(filename):
            df = pd.read_excel(filename)
            st.dataframe(df)
        else:
            st.warning("No data available to show, fill the form first")


    # Streamlit form
    st.title('Data Collection Form')

    with st.form(key='data_form'):
        name = st.text_input('Enter your name')
        choice = st.radio('Gender:', ['Male', 'Female', 'Other'])

        submit_button = st.form_submit_button(label='Submit')

    if st.button('Load Data'):
        display()

    if submit_button:
        if name and choice:
            save_data_to_excel(name, choice)
        else:
            st.error("Please enter your name.")

    if st.button("Logout"):
        cr.key = 0
        st.write('<meta http-equiv="refresh" content="0; url=/login" />', unsafe_allow_html=True)
      # Rerun the app to reflect changes

else:
    st.warning("Access Denied")
    if st.button("Take me to login page"):
        st.write('<meta http-equiv="refresh" content="0; url=/login" />', unsafe_allow_html=True)



